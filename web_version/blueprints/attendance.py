from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from models import AttendanceRecord, Attendance, Student, db
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import io

attendance_bp = Blueprint('attendance', __name__)

@attendance_bp.route('/')
def index():
    """Display all attendance records"""
    records = AttendanceRecord.query.order_by(AttendanceRecord.created_at.desc()).all()
    return render_template('attendance/index.html', records=records)

@attendance_bp.route('/record/<int:record_id>')
def view_record(record_id):
    """View attendance for a specific record"""
    record = AttendanceRecord.query.get_or_404(record_id)
    attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
    
    return render_template('attendance/view_record.html', record=record, attendances=attendances)

@attendance_bp.route('/record/<int:record_id>/update', methods=['POST'])
def update_attendance(record_id):
    """Update attendance status for a student"""
    record = AttendanceRecord.query.get_or_404(record_id)
    student_id = request.form.get('student_id')
    status = request.form.get('status')
    
    if not all([student_id, status]):
        return jsonify({'success': False, 'message': 'Missing required fields'})
    
    # Find or create attendance record
    attendance = Attendance.query.filter_by(record_id=record_id, student_id=student_id).first()
    
    if not attendance:
        # Get student info
        student = Student.query.filter_by(student_id=student_id).first()
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'})
        
        # Create new attendance record
        attendance = Attendance(
            record_id=record_id,
            student_id=student.student_id,
            student_fname=student.fname,
            student_year_level=student.year_level,
            student_course=student.course,
            status=status
        )
        db.session.add(attendance)
    else:
        # Update existing attendance
        attendance.status = status
    
    try:
        # Set timestamp for Present status
        if status == 'Present':
            attendance.timestamp = datetime.now()
        elif status == 'Absent':
            attendance.timestamp = None
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Attendance updated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error updating attendance: {str(e)}'})

@attendance_bp.route('/record/<int:record_id>/initialize', methods=['POST'])
def initialize_record(record_id):
    """Initialize attendance record with all students"""
    record = AttendanceRecord.query.get_or_404(record_id)
    
    # Get all students
    students = Student.query.all()
    
    try:
        for student in students:
            # Check if attendance record already exists
            existing_attendance = Attendance.query.filter_by(
                record_id=record_id, 
                student_id=student.student_id
            ).first()
            
            if not existing_attendance:
                # Create attendance record for this student
                new_attendance = Attendance(
                    record_id=record_id,
                    student_id=student.student_id,
                    student_fname=student.fname,
                    student_year_level=student.year_level,
                    student_course=student.course,
                    status='Absent'
                )
                db.session.add(new_attendance)
        
        db.session.commit()
        flash('Attendance record initialized with all students', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error initializing attendance record: {str(e)}', 'error')
    
    return redirect(url_for('attendance.view_record', record_id=record_id))

@attendance_bp.route('/record/<int:record_id>/mark-present/<student_id>', methods=['POST'])
def mark_present(record_id, student_id):
    """Mark a student as present"""
    attendance = Attendance.query.filter_by(record_id=record_id, student_id=student_id).first_or_404()
    
    try:
        attendance.status = 'Present'
        attendance.timestamp = datetime.now()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Student marked as present'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error marking student as present: {str(e)}'})

@attendance_bp.route('/record/<int:record_id>/mark-absent/<student_id>', methods=['POST'])
def mark_absent(record_id, student_id):
    """Mark a student as absent"""
    attendance = Attendance.query.filter_by(record_id=record_id, student_id=student_id).first_or_404()
    
    try:
        attendance.status = 'Absent'
        attendance.timestamp = None
        db.session.commit()
        return jsonify({'success': True, 'message': 'Student marked as absent'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error marking student as absent: {str(e)}'})

@attendance_bp.route('/record/<int:record_id>/mark-excused/<student_id>', methods=['POST'])
def mark_excused(record_id, student_id):
    """Mark a student as excused"""
    attendance = Attendance.query.filter_by(record_id=record_id, student_id=student_id).first_or_404()
    
    try:
        attendance.status = 'Excused'
        attendance.timestamp = datetime.now()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Student marked as excused'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error marking student as excused: {str(e)}'})

@attendance_bp.route('/record/<int:record_id>/bulk-update', methods=['POST'])
def bulk_update(record_id):
    """Bulk update attendance status"""
    record = AttendanceRecord.query.get_or_404(record_id)
    data = request.get_json()
    status = data.get('status')
    
    if not status:
        return jsonify({'success': False, 'message': 'Status is required'})
    
    try:
        # Get all students
        students = Student.query.all()
        
        for student in students:
            # Find or create attendance record
            attendance = Attendance.query.filter_by(record_id=record_id, student_id=student.student_id).first()
            
            if not attendance:
                # Create new attendance record
                attendance = Attendance(
                    record_id=record_id,
                    student_id=student.student_id,
                    student_fname=student.fname,
                    student_year_level=student.year_level,
                    student_course=student.course,
                    status=status
                )
                db.session.add(attendance)
            else:
                # Update existing attendance
                attendance.status = status
            
            # Set timestamp for Present status
            if status == 'Present':
                attendance.timestamp = datetime.now()
            elif status == 'Absent':
                attendance.timestamp = None
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'All students marked as {status}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error during bulk update: {str(e)}'})

@attendance_bp.route('/record/<int:record_id>/search')
def search_attendance(record_id):
    """Search attendance records"""
    record = AttendanceRecord.query.get_or_404(record_id)
    query = request.args.get('q', '')
    
    if query:
        attendances = Attendance.query.filter(
            Attendance.record_id == record_id,
            (Attendance.student_id.contains(query)) |
            (Attendance.student_fname.contains(query)) |
            (Attendance.student_year_level.contains(query)) |
            (Attendance.student_course.contains(query))
        ).order_by(Attendance.student_id).all()
    else:
        attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
    
    return render_template('attendance/view_record.html', record=record, attendances=attendances, search_query=query)

@attendance_bp.route('/record/<int:record_id>/filter')
def filter_attendance(record_id):
    """Filter attendance records by status"""
    record = AttendanceRecord.query.get_or_404(record_id)
    status_filter = request.args.get('status', 'all')
    
    if status_filter == 'all':
        attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
    else:
        attendances = Attendance.query.filter_by(record_id=record_id, status=status_filter).order_by(Attendance.student_id).all()
    
    return render_template('attendance/view_record.html', record=record, attendances=attendances, status_filter=status_filter)

@attendance_bp.route('/record/<int:record_id>/students')
def get_students_for_record(record_id):
    """Get students for a specific record"""
    record = AttendanceRecord.query.get_or_404(record_id)
    
    # Get all students
    students = Student.query.all()
    
    # Get existing attendance records for this record
    existing_attendances = Attendance.query.filter_by(record_id=record_id).all()
    existing_student_ids = [att.student_id for att in existing_attendances]
    
    # Prepare student data
    student_data = []
    for student in students:
        # Find existing attendance for this student
        existing_attendance = next((att for att in existing_attendances if att.student_id == student.student_id), None)
        
        student_data.append({
            'student_id': student.student_id,
            'student_fname': student.fname,
            'student_year_level': student.year_level,
            'student_course': student.course,
            'status': existing_attendance.status if existing_attendance else 'Not Marked'
        })
    
    return jsonify({'students': student_data})

@attendance_bp.route('/record/<int:record_id>/export-excel')
def export_excel(record_id):
    """Export attendance record to Excel"""
    try:
        record = AttendanceRecord.query.get_or_404(record_id)
        attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
        
        if not attendances:
            flash('No attendance data found for this record', 'warning')
            return redirect(url_for('attendance.view_record', record_id=record_id))
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance - {record.record_name}"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
    
        # Add title
        ws['A1'] = f"Attendance Record: {record.record_name}"
        ws['A1'].font = Font(bold=True, size=16, color="FF6B35")
        ws.merge_cells('A1:G1')
        
        # Add event info
        ws['A2'] = f"Event: {record.event.event_name}"
        ws['A2'].font = Font(bold=True, size=12, color="FF6B35")
        ws.merge_cells('A2:G2')
        
        ws['A3'] = f"Date: {record.created_at.strftime('%B %d, %Y')}"
        ws['A3'].font = Font(bold=True, size=12, color="FF6B35")
        ws.merge_cells('A3:G3')
        
        # Add empty row for spacing
        ws['A4'] = ""
        
        # Add headers
        headers = ['Student ID', 'Name', 'Year Level', 'Course', 'Status', 'Timestamp', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Add data
        for row, attendance in enumerate(attendances, 6):
            ws.cell(row=row, column=1, value=attendance.student_id)
            ws.cell(row=row, column=2, value=attendance.student_fname)
            ws.cell(row=row, column=3, value=attendance.student_year_level)
            ws.cell(row=row, column=4, value=attendance.student_course)
            ws.cell(row=row, column=5, value=attendance.status)
            ws.cell(row=row, column=6, value=attendance.timestamp.strftime('%B %d, %Y %H:%M') if attendance.timestamp and attendance.timestamp is not None else '')
            ws.cell(row=row, column=7, value='')
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            # Check all cells in this column (skip merged cells)
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None and not isinstance(cell, MergedCell):
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            # Set column width (minimum 10, maximum 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary
        summary_row = len(attendances) + 7
        ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=14, color="FF6B35")
        ws.merge_cells(f'A{summary_row}:G{summary_row}')
        
        present_count = sum(1 for a in attendances if a.status == 'Present')
        absent_count = sum(1 for a in attendances if a.status == 'Absent')
        excused_count = sum(1 for a in attendances if a.status == 'Excused')
        total_count = len(attendances)
        
        ws.cell(row=summary_row + 1, column=1, value=f"Total Students: {total_count}").font = Font(bold=True, color="FF6B35")
        ws.cell(row=summary_row + 1, column=3, value=f"Present: {present_count}").font = Font(bold=True, color="008000")
        ws.cell(row=summary_row + 1, column=5, value=f"Absent: {absent_count}").font = Font(bold=True, color="FF0000")
        ws.cell(row=summary_row + 1, column=7, value=f"Excused: {excused_count}").font = Font(bold=True, color="FFA500")
    
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create safe filename by removing/replacing special characters
        safe_record_name = "".join(c for c in record.record_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_record_name = safe_record_name.replace(' ', '_')
        filename = f"attendance_{safe_record_name}_{record.created_at.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Error exporting to Excel: {str(e)}', 'error')
        return redirect(url_for('attendance.view_record', record_id=record_id))
