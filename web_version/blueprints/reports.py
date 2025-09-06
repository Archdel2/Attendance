from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from models import Student, Event, AttendanceRecord, Attendance, db
from datetime import datetime, timedelta
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/')
def index():
    """Reports main page"""
    events = Event.query.order_by(Event.created_at.desc()).all()
    records = AttendanceRecord.query.order_by(AttendanceRecord.created_at.desc()).all()
    
    return render_template('reports/index.html', events=events, records=records)

@reports_bp.route('/event/<int:event_id>')
def event_report(event_id):
    """Generate report for a specific event"""
    event = Event.query.get_or_404(event_id)
    records = AttendanceRecord.query.filter_by(event_id=event_id).all()
    
    # Get attendance data from all records
    attendance_data = []
    for record in records:
        attendances = record.attendances
        for attendance in attendances:
            attendance_data.append({
                'student_id': attendance.student_id,
                'student_fname': attendance.student_fname,
                'student_year_level': attendance.student_year_level,
                'student_course': attendance.course,
                'status': attendance.status,
                'timestamp': attendance.timestamp,
                'record_name': record.record_name
            })
    
    # Calculate statistics
    total_students = len(attendance_data)
    present_count = len([a for a in attendance_data if a['status'] == 'Present'])
    absent_count = len([a for a in attendance_data if a['status'] == 'Absent'])
    excused_count = len([a for a in attendance_data if a['status'] == 'Excused'])
    
    return render_template('reports/event_report.html', 
                         event=event, 
                         attendance_data=attendance_data,
                         total_students=total_students,
                         present_count=present_count,
                         absent_count=absent_count,
                         excused_count=excused_count)

@reports_bp.route('/record/<int:record_id>')
def record_report(record_id):
    """Generate report for a specific attendance record"""
    record = AttendanceRecord.query.get_or_404(record_id)
    attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
    
    # Calculate statistics
    total_students = len(attendances)
    present_count = len([a for a in attendances if a.status == 'Present'])
    absent_count = len([a for a in attendances if a.status == 'Absent'])
    excused_count = len([a for a in attendances if a.status == 'Excused'])
    
    return render_template('reports/record_report.html', 
                         record=record, 
                         attendances=attendances,
                         total_students=total_students,
                         present_count=present_count,
                         absent_count=absent_count,
                         excused_count=excused_count)

@reports_bp.route('/export/event/<int:event_id>/csv')
def export_event_csv(event_id):
    """Export event report to CSV"""
    event = Event.query.get_or_404(event_id)
    records = AttendanceRecord.query.filter_by(event_id=event_id).all()
    
    # Get attendance data
    attendance_data = []
    for record in records:
        attendances = record.attendances
        for attendance in attendances:
            attendance_data.append({
                'student_id': attendance.student_id,
                'student_fname': attendance.student_fname,
                'student_year_level': attendance.student_year_level,
                'student_course': attendance.student_course,
                'status': attendance.status,
                'timestamp': attendance.timestamp,
                'record_name': record.record_name
            })
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Event Report', event.event_name])
    writer.writerow(['Event Date', event.event_date])
    writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    writer.writerow(['Student ID', 'First Name', 'Year Level', 'Course', 'Status', 'Timestamp', 'Record'])
    
    for attendance in attendance_data:
        writer.writerow([
            attendance['student_id'],
            attendance['student_fname'],
            attendance['student_year_level'],
            attendance['student_course'],
            attendance['status'],
            attendance['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if attendance['timestamp'] else '',
            attendance['record_name']
        ])
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'event_report_{event.event_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@reports_bp.route('/export/record/<int:record_id>/csv')
def export_record_csv(record_id):
    """Export record report to CSV"""
    record = AttendanceRecord.query.get_or_404(record_id)
    attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Attendance Record Report', record.record_name])
    writer.writerow(['Event', record.event.event_name])
    writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    writer.writerow(['Student ID', 'First Name', 'Year Level', 'Course', 'Status', 'Timestamp'])
    
    for attendance in attendances:
        writer.writerow([
            attendance.student_id,
            attendance.student_fname,
            attendance.student_year_level,
            attendance.student_course,
            attendance.status,
            attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S') if attendance.timestamp else ''
        ])
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'record_report_{record.record_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@reports_bp.route('/export/event/<int:event_id>/excel')
def export_event_excel(event_id):
    """Export event report to Excel"""
    event = Event.query.get_or_404(event_id)
    records = AttendanceRecord.query.filter_by(event_id=event_id).all()
    
    # Get attendance data
    attendance_data = []
    for record in records:
        attendances = record.attendances
        for attendance in attendances:
            attendance_data.append({
                'student_id': attendance.student_id,
                'student_fname': attendance.student_fname,
                'student_year_level': attendance.student_year_level,
                'student_course': attendance.student_course,
                'status': attendance.status,
                'timestamp': attendance.timestamp,
                'record_name': record.record_name
            })
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event Report"
    
    # Add headers
    ws['A1'] = f"Event Report - {event.event_name}"
    ws['A2'] = f"Event Date: {event.event_date}"
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Total Students: {len(attendance_data)}"
    
    # Style headers
    for cell in ['A1', 'A2', 'A3', 'A4']:
        ws[cell].font = Font(bold=True, size=12)
    
    # Add data headers
    headers = ['Student ID', 'First Name', 'Year Level', 'Course', 'Status', 'Timestamp', 'Record']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    for row, attendance in enumerate(attendance_data, 7):
        ws.cell(row=row, column=1, value=attendance['student_id'])
        ws.cell(row=row, column=2, value=attendance['student_fname'])
        ws.cell(row=row, column=3, value=attendance['student_year_level'])
        ws.cell(row=row, column=4, value=attendance['student_course'])
        ws.cell(row=row, column=5, value=attendance['status'])
        ws.cell(row=row, column=6, value=attendance['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if attendance['timestamp'] else '')
        ws.cell(row=row, column=7, value=attendance['record_name'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'event_report_{event.event_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@reports_bp.route('/export/record/<int:record_id>/excel')
def export_record_excel(record_id):
    """Export record report to Excel"""
    record = AttendanceRecord.query.get_or_404(record_id)
    attendances = Attendance.query.filter_by(record_id=record_id).order_by(Attendance.student_id).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Record Report"
    
    # Add headers
    ws['A1'] = f"Attendance Record Report - {record.record_name}"
    ws['A2'] = f"Event: {record.event.event_name}"
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Total Students: {len(attendances)}"
    
    # Style headers
    for cell in ['A1', 'A2', 'A3', 'A4']:
        ws[cell].font = Font(bold=True, size=12)
    
    # Add data headers
    headers = ['Student ID', 'First Name', 'Year Level', 'Course', 'Status', 'Timestamp']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data
    for row, attendance in enumerate(attendances, 7):
        ws.cell(row=row, column=1, value=attendance.student_id)
        ws.cell(row=row, column=2, value=attendance.student_fname)
        ws.cell(row=row, column=3, value=attendance.student_year_level)
        ws.cell(row=row, column=4, value=attendance.student_course)
        ws.cell(row=row, column=5, value=attendance.status)
        ws.cell(row=row, column=6, value=attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S') if attendance.timestamp else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'record_report_{record.record_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@reports_bp.route('/summary')
def summary():
    """Generate summary report"""
    # Get overall statistics
    total_students = Student.query.count()
    total_events = Event.query.count()
    total_records = AttendanceRecord.query.count()
    
    # Get recent attendance statistics
    today = datetime.now().date()
    today_attendance = Attendance.query.filter(
        Attendance.timestamp >= today
    ).count()
    
    # Get attendance by status
    present_count = Attendance.query.filter_by(status='Present').count()
    absent_count = Attendance.query.filter_by(status='Absent').count()
    excused_count = Attendance.query.filter_by(status='Excused').count()
    
    # Get recent events
    recent_events = Event.query.order_by(Event.created_at.desc()).limit(10).all()
    
    return render_template('reports/summary.html',
                         total_students=total_students,
                         total_events=total_events,
                         total_records=total_records,
                         today_attendance=today_attendance,
                         present_count=present_count,
                         absent_count=absent_count,
                         excused_count=excused_count,
                         recent_events=recent_events)
